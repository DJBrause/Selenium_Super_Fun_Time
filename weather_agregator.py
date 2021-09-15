from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook
import datetime
import time
import random
from selenium.webdriver.chrome.options import Options

option = Options()
option.add_argument("--disable-infobars")
weather_com = "https://weather.com/"
accuweather_com = "https://www.accuweather.com/"
pogoda_onet = "https://pogoda.onet.pl/"
twoja_pogoda = "https://www.twojapogoda.pl/"
meteoprog = "https://www.meteoprog.pl/pl/"
header = ["Today", "Tomorrow", "In 2 days", "In 3 days"]
date_ = datetime.datetime.now()


print("Enter city name: ")
searched_phrase = input()
wb = Workbook()
dest_filename = f'{searched_phrase.lower()}_weather_forecast_{date_.day}_{date_.month}_{date_.year}.xlsx'

site_list = [weather_com, accuweather_com, pogoda_onet, twoja_pogoda, meteoprog]

PATH = "chromedriver_win32/chromedriver.exe"
driver = webdriver.Chrome(PATH)


def weather_com_fun():
    print(weather_com)

    weather = []
    temperature = []
    ws = wb.active
    ws.title = "Weather.com"

    driver.get(weather_com)
    driver.maximize_window()
    time.sleep(random.uniform(2, 3))
    cookies_but = driver.find_element_by_class_name("truste-button2")
    cookies_but.click()
    time.sleep(random.uniform(1, 4))
    search_bar = driver.find_element_by_id("LocationSearch_input")
    search_bar.send_keys(str(searched_phrase))
    time.sleep(random.uniform(1, 3))
    search_bar.send_keys(Keys.RETURN)
    time.sleep(random.uniform(0.5, 1.2))
    weather_today = driver.find_element_by_css_selector('[data-testid="wxPhrase"]')
    weather.append(str(weather_today.text))
    temperature_forecast = driver.find_element_by_css_selector('[data-from-string="localsuiteNav_3_10-dniowa"]')
    temperature_forecast.click()
    weather_tomorrow = driver.find_elements_by_css_selector('[data-testid="TemperatureValue"]')
    str_weather_tomorrow = []
    str_ten_days_forecast = []
    for item in weather_tomorrow:
        if item.text != '':
            str_weather_tomorrow.append(item.text)
    temperature.append(str_weather_tomorrow[0])
    temperature.append(str_weather_tomorrow[2])
    temperature.append(str_weather_tomorrow[4])
    temperature.append(str_weather_tomorrow[6])
    ten_days_forecast_weather = driver.find_elements_by_class_name('DetailsSummary--extendedData--365A_')
    for item in ten_days_forecast_weather:
        if item.text != '':
            str_ten_days_forecast.append(item.text)

    weather.append(str_ten_days_forecast[0])
    weather.append(str_ten_days_forecast[2])
    weather.append(str_ten_days_forecast[4])

    ws.append(header)
    ws.append(weather)
    ws.append(temperature)


def accuweather_fun():
    print(accuweather_com)

    weather = []
    temperature = []
    ws = wb.create_sheet("Accuweather.com")
    driver.get(accuweather_com)
    driver.maximize_window()
    time.sleep(random.uniform(2, 3))
    cookies_but = driver.find_element_by_css_selector('[aria-label="Consent"]')
    cookies_but.click()
    time.sleep(random.uniform(1, 4))
    search_bar = driver.find_element_by_class_name("search-input")
    search_bar.send_keys(str(searched_phrase))
    time.sleep(random.uniform(1, 3))
    search_bar.send_keys(Keys.RETURN)
    time.sleep(random.uniform(1.5, 2.2))
    results = driver.find_elements_by_partial_link_text(searched_phrase)
    try:
        results[0].click()
    except IndexError:
        pass
    time.sleep(random.randint(2, 3))

    close_but = driver.find_element_by_class_name("header-city-link")
    driver.execute_script("arguments[0].click();", close_but)  # bypasses google vignette
    time.sleep(random.uniform(1.5, 2.2))
    daily = driver.find_element_by_css_selector('[data-qa="daily"]')
    daily.click()

    temp_ = driver.find_elements_by_xpath('//span[@class="high"]')
    for i in range(4):
        temperature.append(temp_[i].text)

    weather_ = driver.find_elements_by_xpath('//div[@class="phrase"]')
    for i in range(4):
        weather.append(weather_[i].text)

    driver.quit()
    ws.append(header)
    ws.append(weather)
    ws.append(temperature)


def avarage_values_fun():
    ws_avg = wb.create_sheet("Avarage Values")
    ws1 = wb["Weather.com"]
    ws2 = wb["Accuweather.com"]
    row_ws1 = ws1[3]
    row_ws2 = ws2[3]
    int_row_ws1 = []
    int_row_ws2 = []

    for cell in row_ws1:
        try:
            x = str(cell.value)
            x = x.strip("°")
            int_row_ws1.append(x)
        except ValueError:
            pass
    for cell in row_ws2:
        try:
            y = str(cell.value)
            y = y.strip("°")
            int_row_ws2.append(y)
        except ValueError:
            pass

    avg_values = []
    for val in range(4):
        avg_values.append(str((int(int_row_ws1[val]) + int(int_row_ws2[val]))/2)+"°")

    ws_avg.append(header)
    ws_avg.append(avg_values)


weather_com_fun()
accuweather_fun()
avarage_values_fun()
wb.save(filename=dest_filename)

print(f"Task complete! File name is {dest_filename}")
